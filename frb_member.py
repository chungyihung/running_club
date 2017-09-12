#!/usr/bin/python3
# -*- coding:utf-8 -*-

import pyrebase as pybs
import json
import collections as clt
import openpyxl as pyxl
import enum as enm
import datetime

import frb_config as frbcfg     # This should be added manually

''' ---------------------------------------------
Member Key
----------------------------------------------'''
MemberID    = "id"
Position    = "position"
Name        = "name"
IDCard      = "idcard"
BirthY      = "birth_Y"
BirthM      = "birth_M"
BirthD      = "birth_D"
Area        = "area"
CellPhone   = "cellphone"
Phone       = "phone"
Phone2      = "phone2"
Address     = "address"
Photo       = "photo"
TShirt_SZ   = "tshirt_sz"
Coat_SZ     = "coat_sz"

''' ---------------------------------------------
Default Member attributes
----------------------------------------------'''
DEFAULT_PHOTO = "Resource/none_photo.jpg"
DEFAULT_TSHIRT_SZ = 0
DEFAULT_COAT_SZ = 0

INVALID_ID = 65535

''' ---------------------------------------------
Excel file related
----------------------------------------------'''
DB_FILE_NAME            = "running_club_member.db"
EXCEL_FILE_NAME         = "running_table.xlsx"
EXCEL_FILE_TS_NAME      = "running_table_test.xlsx"
EXCEL_SHEET_ALL_NAME    = "新會員(全)"

''' ---------------------------------------------
Excel items sequence
----------------------------------------------'''
class ExcelItem( enm.IntEnum ):
    EXSL_POSITION           = 1
    EXSL_NAME               = 2
    EXSL_IDCARD             = 3
    EXSL_BIRTHDAY_Y_STR     = 4
    EXSL_BIRTHDAY_STR       = 5
    EXSL_AREA               = 6
    EXSL_CELLPHONE          = 7
    EXSL_PHONE              = 8
    EXSL_PHONE2             = 9
    EXSL_ADDRESS            = 10
    EXSL_CNT                = enm.auto()

EXSL_BIRTH_DELIM        = '.'

class frb_member:
    def __init__( self ):
        self.firebase = pybs.initialize_app(frbcfg.config)
        self.db = self.firebase.database()
        self.stg = self.firebase.storage()

    ''' ---------------------------------------------
    Get member data in dict type
    ----------------------------------------------'''
    def get_member( self, memid ):
        resp = self.db.child("users").child(memid).get()
        if resp.val() != None:
            return dict(resp.val())
        else:
            print("No {} member data in FRB server".format(memid))
            return None

    ''' ---------------------------------------------
    Set member data with dict type
    ----------------------------------------------'''
    def set_member( self, data, memid ):
        self.db.child("users").child(memid).set(data)
        self.updt_timestamp()

    ''' ---------------------------------------------
    Update member data with dict type
    TODO: 1. Implement item specify feature
    ----------------------------------------------'''
    def upt_member( self, data, memid ):
        self.db.child("users").child(memid).update(data)
        self.updt_timestamp()

    ''' ---------------------------------------------
    Delete member
    ----------------------------------------------'''
    def del_member( self, memid ):
        self.db.child("users").child(memid).remove()
        self.updt_timestamp()

    ''' ---------------------------------------------
    Get all member data in list type
    ----------------------------------------------'''
    def get_all_member( self ):
        resp = self.db.child("users").get()
        if resp.val() != None:
            return resp.val()
        else:
            print("No member data in FRB server")
            return None

    ''' ---------------------------------------------
    Time stamp
    ----------------------------------------------'''
    def updt_timestamp( self ):
        self.db.child("timestamp").set(int(datetime.datetime.utcnow().timestamp()))

    def get_timestamp( self ):
        timestamp = self.db.child("timestamp").get()
        if timestamp.val() != None:
            return timestamp.val()
        else:
            return 0

    def compare_timestamp( self, local_t ):
        server_t = self.get_timestamp()
        return local_t - server_t

    def init_member( self, memid = INVALID_ID ):
        init_member = {}
        init_member[MemberID] = memid
        init_member[Position] = ""
        init_member[Name] = ""
        init_member[IDCard] = ""
        init_member[BirthY] = 0
        init_member[BirthM] = 0
        init_member[BirthD] = 0
        init_member[Area] = ""
        init_member[CellPhone] = ""
        init_member[Phone] = ""
        init_member[Phone2] = ""
        init_member[Address] = ""
        init_member[Photo] = DEFAULT_PHOTO
        init_member[TShirt_SZ] = DEFAULT_TSHIRT_SZ
        init_member[Coat_SZ] = DEFAULT_COAT_SZ
        return init_member

    ''' ---------------------------------------------
    External Utility
    ----------------------------------------------'''
    def birthday_str_to_list( self, birth_str ):
        birth_list = [ 0, 1, 1 ]
        try:
            birth_list = birth_str.split( EXSL_BIRTH_DELIM )
        except:
            print( "Convert birthday string to list failed, using default value" )
            birth_list = [ 0, 1, 1 ]

        return birth_list

    def xstr( self, s ):
        if s is None:
            return ''
        return str(s)

    ''' ---------------------------------------------
    Database <--> Excel file convertion
    ----------------------------------------------'''
    def cnvt_excel_to_db( self ):
        """
        Load excel file and specific the sheet
        """
        wb = pyxl.load_workbook( EXCEL_FILE_NAME )
        sheet_all = wb.get_sheet_by_name( EXCEL_SHEET_ALL_NAME )

        """
        Initialize
        """
        add_id = 1

        for item in sheet_all.iter_rows( row_offset = 1 ):
            if item[0].value != None:
                print(int(item[0].value))
                mbr_birthday_lst = self.birthday_str_to_list( item[ExcelItem.EXSL_BIRTHDAY_STR].value )

                user_info = {}
                user_info[MemberID] = int(add_id)
                user_info[Position] = self.xstr( item[ExcelItem.EXSL_POSITION].value )
                user_info[Name] = self.xstr( item[ExcelItem.EXSL_NAME].value )
                user_info[IDCard] = self.xstr( item[ExcelItem.EXSL_IDCARD].value )
                user_info[BirthY] = int(mbr_birthday_lst[0])
                user_info[BirthM] = int(mbr_birthday_lst[1])
                user_info[BirthD] = int(mbr_birthday_lst[2])
                user_info[Area] = self.xstr( item[ExcelItem.EXSL_AREA].value )
                user_info[CellPhone] = self.xstr( item[ExcelItem.EXSL_CELLPHONE].value )
                user_info[Phone] = self.xstr( item[ExcelItem.EXSL_PHONE].value )
                user_info[Phone2] = self.xstr( item[ExcelItem.EXSL_PHONE2].value )
                user_info[Address] = self.xstr( item[ExcelItem.EXSL_ADDRESS].value )
                user_info[Photo] = DEFAULT_PHOTO
                user_info[TShirt_SZ] = DEFAULT_TSHIRT_SZ
                user_info[Coat_SZ] = DEFAULT_COAT_SZ

                self.set_member(user_info, add_id)

                add_id += 1
                if add_id % 10 == 4:
                    add_id += 1

    def cnvt_db_to_excel( self ):
        wb = pyxl.Workbook()
        for sheet in wb.worksheets:
            wb.remove_sheet( sheet )

        ws = wb.create_sheet( title=EXCEL_SHEET_ALL_NAME )
        ws.append( ["編號", "職稱", "姓名", "身分證", "年次", "出生年月日", "地址", "手機", "電話1", "電話2", "通訊處(地址)"] )

        member = self.get_all_member()

        for mem in member:
            if mem != None:
                append_list = [ mem["id"], mem["position"], mem["name"],mem["idcard"],mem["birth_Y"],
                                "{}{}{}{}{}".format( mem["birth_Y"], EXSL_BIRTH_DELIM, mem["birth_M"], EXSL_BIRTH_DELIM, mem["birth_D"] ),
                                mem["address"], mem["cellphone"], mem["phone"], mem["phone2"], mem["address"]]
                #print( append_list )
                ws.append( append_list )

        wb.save( filename = EXCEL_FILE_TS_NAME )

