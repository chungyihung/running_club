#!/usr/bin/python3
# -*- coding:utf-8 -*-

import sqlite3 as sql3
import openpyxl as pyxl
import enum as enm

INVALID_ID = 65535
INVALID_CARD_ID = '##########'
INVALID_PHONE = '09__-______'

DEFAULT_PHOTO = "Resource/none_photo.jpg"
DEFAULT_TSHIRT_SZ = 0
DEFAULT_COAT_SZ = 0

''' ---------------------------------------------
Excel file related
----------------------------------------------'''
DB_FILE_NAME            = "running_club_member.db"
EXCEL_FILE_NAME         = "running_table.xlsx"
EXCEL_FILE_TS_NAME      = "running_table_test.xlsx"
EXCEL_SHEET_ALL_NAME    = "新會員(全)"

''' ---------------------------------------------
Excel items sequence
----------------------------------------------'''
class ExcelItem( enm.IntEnum ):
    EXSL_POSITION           = 1
    EXSL_NAME               = 2
    EXSL_IDCARD             = 3
    EXSL_BIRTHDAY_Y_STR     = 4
    EXSL_BIRTHDAY_STR       = 5
    EXSL_AREA               = 6
    EXSL_CELLPHONE          = 7
    EXSL_PHONE              = 8
    EXSL_PHONE2             = 9
    EXSL_ADDRESS            = 10
    EXSL_CNT                = enm.auto()

EXSL_BIRTH_DELIM        = '.'
''' ---------------------------------------------
Excel items sequence
----------------------------------------------'''
class DBItem( enm.IntEnum ):
    DB_id               = 0
    DB_position         = 1
    DB_name             = 2
    DB_idcard           = 3
    DB_birthday_Y       = 4
    DB_birthday_M       = 5
    DB_birthday_D       = 6
    DB_area             = 7
    DB_cell_phone       = 8
    DB_phone            = 9
    DB_phone2           = 10
    DB_address          = 11
    DB_photo            = 12
    DB_tshirt_sz        = 13
    DB_coat_sz          = 14
    DB_ITEM_CNT         = enm.auto()

''' ---------------------------------------------
Main member class
----------------------------------------------'''
class member:
    def __init__(self, **meminfo):
        self.__name         = meminfo.get( 'name', 'Unknown'            )       # String
        self.__mem_id       = meminfo.get( 'id', INVALID_ID             )       # Unsigned Int
        self.__cell_phone   = meminfo.get( 'cell_phone', INVALID_PHONE  )       # String
        self.__position     = meminfo.get( 'position', 'Unknown'        )       # String. Job position
        self.__id_card      = meminfo.get( 'idcard', INVALID_CARD_ID    )       # String
        self.__birthday_Y   = meminfo.get( 'birthday_Y', ''             )       # Unsigned Int in Republic Era.
        self.__birthday_M   = meminfo.get( 'birthday_M', ''             )       # Unsigned Int
        self.__birthday_D   = meminfo.get( 'birthday_D', ''             )       # Unsigned Int
        self.__area         = meminfo.get( 'area', ''                   )       # String. (Using enum to represent)
        self.__address      = meminfo.get( 'address', ''                )       # String
        self.__phone        = meminfo.get( 'phone', ''                  )       # String
        self.__phone2       = meminfo.get( 'phone2', ''                 )       # String
        self.__photo        = meminfo.get( 'photo', ''                  )       # String. Recording relative file path
        self.__tshirt_sz    = meminfo.get( 'tshirt_sz', ''              )       # int
        self.__coat_sz      = meminfo.get( 'coat_sz', ''                )       # int

        '''
        Create database if not exist
        '''
        self.create_tbl()

    @property
    def name(self):
        return self.__name

    @property
    def mem_id(self):
        return self.__mem_id

    @property
    def cell_phone(self):
        return self.__cell_phone

    @property
    def position(self):
        return self.__position

    @property
    def id_card(self):
        return self.__id_card

    @property
    def birthday_Y(self):
        return self.__birthday_Y

    @property
    def birthday_M(self):
        return self.__birthday_M

    @property
    def birthday_D(self):
        return self.__birthday_D

    @property
    def area(self):
        return self.__area

    @property
    def address(self):
        return self.__address

    @property
    def phone(self):
        return self.__phone

    @property
    def phone2(self):
        return self.__phone2

    @property
    def photo(self):
        return self.__photo

    @property
    def tshirt_sz(self):
        return self.__tshirt_sz

    @property
    def coat_sz(self):
        return self.__coat_sz

    @name.setter
    def name(self, name):
        self.__name = name

    @mem_id.setter
    def mem_id(self, mem_id):
        self.__mem_id = mem_id

    @cell_phone.setter
    def cell_phone(self, cell_phone_number):
        self.__cell_phone = cell_phone_number

    @position.setter
    def position(self, position):
        self.__position = position

    @id_card.setter
    def id_card(self, id_card):
        self.__id_card = id_card

    @birthday_Y.setter
    def birthday_Y(self, birthday_Y):
        self.__birthday_Y = birthday_Y

    @birthday_M.setter
    def birthday_M(self, birthday_M):
        self.__birthday_M = birthday_M

    @birthday_D.setter
    def birthday_D(self, birthday_D):
        self.__birthday_D = birthday_D

    @area.setter
    def area(self, area):
        self.__area = area

    @address.setter
    def address(self, address):
        self.__address = address

    @phone.setter
    def phone(self, phone):
        self.__phone = phone

    @phone2.setter
    def phone2(self, phone2):
        self.__phone2 = phone2

    @photo.setter
    def photo(self, photo):
        self.__photo = photo

    @tshirt_sz.setter
    def tshirt_sz(self, tshirt_sz):
        self.__tshirt_sz = tshirt_sz

    @coat_sz.setter
    def coat_sz(self, coat_sz):
        self.__coat_sz = coat_sz

    '''
    Member functions
    '''
    def xstr( self, s ):
        if s is None:
            return ''
        return str(s)

    def create_tbl( self ):
        with sql3.connect( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute( ''' CREATE TABLE IF NOT EXISTS member (
                        id          INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
                        position    TEXT NOT NULL,
                        name        TEXT NOT NULL,
                        idcard      TEXT,
                        birthday_Y  INTEGER,
                        birthday_M  INTEGER,
                        birthday_D  INTEGER,
                        area        TEXT,
                        cell_phone  TEXT,
                        phone       TEXT,
                        phone2      TEXT,
                        address     TEXT,
                        photo       TEXT,
                        tshirt_sz   INTEGER,
                        coat_sz     INTEGER
                        )''' )

    def drop_tbl( self ):
        with sql3.connect( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute( "DROP TABLE member")

    def load_from_db( self, mem_id ):
        with sql3.connect( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute( "SELECT * FROM member WHERE id=?", (mem_id, ) )
            result = c.fetchone()
            if result != None:
                self.__mem_id       = result[0]
                self.__position     = result[1]
                self.__name         = result[2]
                self.__id_card      = result[3]
                self.__birthday_Y   = result[4]
                self.__birthday_M   = result[5]
                self.__birthday_D   = result[6]
                self.__area         = result[7]
                self.__cell_phone   = result[8]
                self.__phone        = result[9]
                self.__phone2       = result[10]
                self.__address      = result[11]
                self.__photo        = result[12]
                self.__tshirt_sz    = result[13]
                self.__coat_sz      = result[14]
                print("Now current member ID is {}".format(self.__mem_id))
            else:
                print("Fetch nothing in DB")

    def save_to_db( self ):

        with sql3.connect( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute( '''SELECT MAX(id) FROM member''' )

            max_id = c.fetchone()[0]
            if max_id == None:
                max_id = 0
            else:
                max_id += 1
                if max_id % 10 == 4:
                    max_id += 1

            self.__mem_id = max_id
            c.execute( '''INSERT INTO member
                          ( id, position, name, idcard, birthday_Y, birthday_M, birthday_D, area, cell_phone, phone, phone2, address, photo, tshirt_sz, coat_sz )
                          VALUES ( ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ? )''',
                          ( self.__mem_id,
                            self.__position,
                            self.__name,
                            self.__id_card,
                            self.__birthday_Y,
                            self.__birthday_M,
                            self.__birthday_D,
                            self.__area,
                            self.__cell_phone,
                            self.__phone,
                            self.__phone2,
                            self.__address,
                            self.__photo,
                            self.__tshirt_sz,
                            self.__coat_sz
                            ) )

    def update_to_db( self ):
        with sql3.connect( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute( '''UPDATE member SET
                      position=:position,
                      name=:name,
                      idcard=:idcard,
                      birthday_Y=:birthday_Y,
                      birthday_M=:birthday_M,
                      birthday_D=:birthday_D,
                      area=:area,
                      cell_phone=:cell_phone,
                      phone=:phone,
                      phone2=:phone2,
                      address=:address,
                      photo=:photo,
                      tshirt_sz=:tshirt_sz,
                      coat_sz=:coat_sz
                      WHERE id =:id ''',
                      { "position"      : self.__position,
                        "name"          : self.__name,
                        "idcard"        : self.__id_card,
                        "birthday_Y"    : self.__birthday_Y,
                        "birthday_M"    : self.__birthday_M,
                        "birthday_D"    : self.__birthday_D,
                        "area"          : self.__area,
                        "cell_phone"    : self.__cell_phone,
                        "phone"         : self.__phone,
                        "phone2"        : self.__phone2,
                        "address"       : self.__address,
                        "photo"         : self.__photo,
                        "id"            : self.__mem_id,
                        "tshirt_sz"     : self.__tshirt_sz,
                        "coat_sz"       : self.__coat_sz
                        } )

    def delete_item( self, del_id ):
        with sql3.connect( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute( '''DELETE FROM member WHERE id = ?''', ( del_id, ) )

    def birthday_str_to_list( self, birth_str ):
        birth_list = [ 0, 1, 1 ]
        try:
            birth_list = birth_str.split( EXSL_BIRTH_DELIM )
        except:
            print( "Convert birthday string to list failed, using default value" )
            birth_list = [ 0, 1, 1 ]

        return birth_list

    def retrieve_all_data( self ):
        with sql3.Connection( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute('''SELECT * FROM member ''')
            rows = c.fetchall()
            return rows

    ''' ---------------------------------------------
    Database <--> Excel file convertion
    ----------------------------------------------'''
    def cnvt_excel_to_db( self ):
        """
        Load excel file and specific the sheet
        """
        wb = pyxl.load_workbook( EXCEL_FILE_NAME )
        sheet_all = wb.get_sheet_by_name( EXCEL_SHEET_ALL_NAME )

        """
        Delete and recreate the table of database
        """
        self.drop_tbl()
        self.create_tbl()

        add_id = 1

        with sql3.Connection( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            for item in sheet_all.iter_rows( row_offset = 1 ):
                if item[0].value != None:
                    print(int(item[0].value))
                    mbr_birthday_lst = self.birthday_str_to_list( item[ExcelItem.EXSL_BIRTHDAY_STR].value )
                    c.execute( '''INSERT INTO member
                                  ( id, position, name, idcard, birthday_Y, birthday_M, birthday_D, area, cell_phone, phone, phone2, address, photo, tshirt_sz, coat_sz )
                                  VALUES ( ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ? )''',
                                  ( add_id,
                                    self.xstr( item[ExcelItem.EXSL_POSITION].value ),
                                    self.xstr( item[ExcelItem.EXSL_NAME].value ),
                                    self.xstr( item[ExcelItem.EXSL_IDCARD].value ),
                                    mbr_birthday_lst[0],
                                    mbr_birthday_lst[1],
                                    mbr_birthday_lst[2],
                                    self.xstr( item[ExcelItem.EXSL_AREA].value ),
                                    self.xstr( item[ExcelItem.EXSL_CELLPHONE].value ),
                                    self.xstr( item[ExcelItem.EXSL_PHONE].value ),
                                    self.xstr( item[ExcelItem.EXSL_PHONE2].value ),
                                    self.xstr( item[ExcelItem.EXSL_ADDRESS].value ),
                                    DEFAULT_PHOTO,
                                    DEFAULT_TSHIRT_SZ,
                                    DEFAULT_COAT_SZ ) )
                    add_id += 1
                    if add_id % 10 == 4:
                        add_id += 1

    def cnvt_db_to_excel( self ):
        wb = pyxl.Workbook()
        for sheet in wb.worksheets:
            wb.remove_sheet( sheet )

        ws = wb.create_sheet( title=EXCEL_SHEET_ALL_NAME )
        ws.append( ["編號", "職稱", "姓名", "身分證", "年次", "出生年月日", "地址", "手機", "電話1", "電話2", "通訊處(地址)"] )
        with sql3.Connection( DB_FILE_NAME ) as conn:
            c = conn.cursor()
            c.execute( "SELECT * FROM member" )
            for row in c:
                append_list = [ row[0], row[1],row[2],row[3],row[4],
                               "{}{}{}{}{}".format( row[4], EXSL_BIRTH_DELIM, row[5], EXSL_BIRTH_DELIM, row[6] ),
                                row[7], row[8], row[9], row[10], row[11]]
                #print( append_list )
                ws.append( append_list )

        wb.save( filename = EXCEL_FILE_TS_NAME )
